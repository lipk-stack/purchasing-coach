"""Web UI endpoints exercised over real HTTP with the fake LLM backend."""

import json
import threading
import urllib.error
import urllib.request

import pytest
from openpyxl import load_workbook

import coach.webui as webui_mod
from coach.llm import Coach
from coach.webui import WebUI
from tests.test_tender import CHECKLIST, FakeBackend


@pytest.fixture()
def server(tmp_path):
    ui = WebUI(Coach("guideline text", FakeBackend()), FakeBackend(),
               "guideline.docx", None, tmp_path)
    httpd = ui.make_server(port=0)
    thread = threading.Thread(target=httpd.serve_forever, daemon=True)
    thread.start()
    base = f"http://127.0.0.1:{httpd.server_address[1]}"
    yield base, ui
    httpd.shutdown()


def _get(url):
    with urllib.request.urlopen(url, timeout=10) as resp:
        return resp.status, resp.headers, resp.read()


def _post(url, payload):
    req = urllib.request.Request(
        url, data=json.dumps(payload).encode(),
        headers={"Content-Type": "application/json"}, method="POST")
    with urllib.request.urlopen(req, timeout=10) as resp:
        return resp.status, json.loads(resp.read())


def test_index_and_meta(server):
    base, _ = server
    status, headers, body = _get(base + "/")
    assert status == 200 and b"Purchasing Coach" in body
    assert "text/html" in headers["Content-Type"]

    status, _, body = _get(base + "/api/meta")
    meta = json.loads(body)
    assert status == 200
    assert meta["backend"] == "fake"
    assert meta["model"] == "fake-model"
    assert meta["guideline"] == "guideline.docx"


def test_page_has_restart_interview_wiring(server):
    base, _ = server
    _, _, body = _get(base + "/")
    page = body.decode()
    # Mid-interview the tender button becomes a restart control, and the
    # typed "restart" / "/tender" inputs re-enter startTender().
    assert "Restart Interview" in page or "Restart interview" in page
    assert "function endTender()" in page
    assert "'restart'" in page and "'/tender'" in page


def test_page_meets_accessibility_contract(server):
    """The served SPA carries the WCAG 2.2 AA landmarks/ARIA we rely on."""
    base, _ = server
    _, _, body = _get(base + "/")
    page = body.decode()
    # Document language + a single top-level heading.
    assert '<html lang="en">' in page
    assert page.count("<h1") == 1
    # Bypass-blocks: a skip link targeting the main landmark.
    assert 'class="skip-link" href="#main"' in page
    assert '<main class="main" id="main"' in page
    # Primary nav items are real buttons (keyboard operable), not click divs,
    # and the active one is marked for assistive tech.
    assert '<button type="button" class="nav-item active"' in page
    assert 'aria-current="page"' in page
    # No legacy click-only <div> nav items remain.
    assert '<div class="nav-item' not in page
    # Stateful controls expose their state.
    assert 'aria-pressed=' in page and 'aria-expanded=' in page
    # Form controls in the checklist toolbar are labelled.
    assert 'for="clSearch"' in page and 'role="search"' in page
    # Canvas charts have a text alternative for screen readers.
    assert 'role="img" aria-labelledby="pieTitle"' in page
    assert 'id="pieDesc"' in page and 'id="barDesc"' in page
    # Keyboard alternative to drag-and-drop reordering (WCAG 2.1.1).
    assert "function moveRow(" in page
    assert "ArrowUp" in page and "ArrowDown" in page


def test_chat_streams_reply(server):
    base, _ = server
    req = urllib.request.Request(
        base + "/api/chat",
        data=json.dumps({"messages": [{"role": "user",
                                       "content": "warranty?"}]}).encode(),
        headers={"Content-Type": "application/json"}, method="POST")
    with urllib.request.urlopen(req, timeout=10) as resp:
        assert resp.status == 200
        text = resp.read().decode()
    assert text == "See clause 8.4 (Warranty)."


def test_tender_flow_over_http(server, tmp_path):
    base, ui = server
    status, plan = _post(base + "/api/tender/start", {"item": "Firewalls"})
    assert status == 200
    assert [q["question"] for q in plan["questions"]] == [
        "When is the submission deadline?",
        "Which department is requesting this?"]

    answers = [[q["question"], "answer"] for q in plan["questions"]]
    status, result = _post(base + "/api/tender/finish",
                           {"item": "Firewalls", "answers": answers})
    assert status == 200
    assert result["count"] == len(CHECKLIST["requirements"])
    assert result["file"].startswith("TENDER_CHECKLIST_Firewall_appliances")
    assert (ui.out_dir / result["file"]).exists()

    status, headers, body = _get(base + result["download"])
    assert status == 200
    assert "spreadsheetml" in headers["Content-Type"]
    download = tmp_path / "downloaded.xlsx"
    download.write_bytes(body)
    wb = load_workbook(download)
    assert "Compliance Tracker" in wb.sheetnames


def test_bad_requests(server):
    base, _ = server
    for url, payload in [(base + "/api/tender/start", {}),
                         (base + "/api/chat", {})]:
        with pytest.raises(urllib.error.HTTPError) as err:
            _post(url, payload)
        assert err.value.code == 400

    # Downloads are limited to files generated in this session.
    with pytest.raises(urllib.error.HTTPError) as err:
        _get(base + "/api/download/../../etc/passwd")
    assert err.value.code == 404
    with pytest.raises(urllib.error.HTTPError) as err:
        _get(base + "/api/download/unknown.xlsx")
    assert err.value.code == 404


class _FailingBackend(FakeBackend):
    def complete_json(self, *a, **k):
        raise RuntimeError("boom")


def test_post_error_returns_500_and_is_logged(tmp_path, caplog):
    ui = WebUI(Coach("guideline text", _FailingBackend()), _FailingBackend(),
               "g.docx", None, tmp_path)
    httpd = ui.make_server(port=0)
    thread = threading.Thread(target=httpd.serve_forever, daemon=True)
    thread.start()
    base = f"http://127.0.0.1:{httpd.server_address[1]}"
    try:
        with caplog.at_level("ERROR", logger="coach.webui"):
            with pytest.raises(urllib.error.HTTPError) as err:
                _post(base + "/api/tender/start", {"item": "x"})
            assert err.value.code == 500
        assert any("failed" in r.getMessage() for r in caplog.records)
    finally:
        httpd.shutdown()


def test_security_headers_present(server):
    base, _ = server
    _, headers, _ = _get(base + "/api/meta")
    assert headers.get("X-Content-Type-Options") == "nosniff"


def test_oversized_body_rejected(server, monkeypatch):
    base, _ = server
    monkeypatch.setattr(webui_mod, "MAX_BODY_BYTES", 8)
    with pytest.raises(urllib.error.HTTPError) as err:
        _post(base + "/api/chat",
              {"messages": [{"role": "user", "content": "hello world"}]})
    assert err.value.code == 413


def test_foreign_host_header_is_rejected(server):
    """DNS-rebinding defence: only loopback Host headers are served."""
    base, _ = server
    # A spoofed (attacker-controlled) Host is refused with 403...
    req = urllib.request.Request(
        base + "/api/meta", headers={"Host": "evil.example.com"})
    with pytest.raises(urllib.error.HTTPError) as err:
        urllib.request.urlopen(req, timeout=10)
    assert err.value.code == 403

    # ...while an explicit loopback Host (with port) is accepted.
    port = base.rsplit(":", 1)[1]
    req = urllib.request.Request(
        base + "/api/meta", headers={"Host": f"localhost:{port}"})
    with urllib.request.urlopen(req, timeout=10) as resp:
        assert resp.status == 200


def test_session_id_traversal_is_neutralised(tmp_path, monkeypatch):
    sdir = tmp_path / "sessions"
    sdir.mkdir()
    monkeypatch.setattr(webui_mod, "SESSIONS_DIR", sdir)
    ui = WebUI(Coach("g", FakeBackend()), FakeBackend(), "g.docx", None, tmp_path)

    # A malicious id is not honoured — a safe id is minted and the file stays
    # inside the sessions directory.
    sid = ui.save_session({"id": "../../evil", "title": "x"})
    assert sid != "../../evil"
    assert (sdir / f"{sid}.json").exists()
    assert not (tmp_path / "evil.json").exists()

    # Reads/deletes with traversal ids are safe no-ops.
    assert ui._session_path("../bad") is None
    assert ui.load_session("../../evil") is None
    assert ui.delete_session("../../etc/passwd") is False


def test_normalize_history_coerces_untrusted_input():
    # Well-formed messages pass through unchanged.
    assert webui_mod._normalize_history(
        [{"role": "user", "content": "hi"}]) == [{"role": "user", "content": "hi"}]
    # Items that would crash a retrieval backend's messages[-1]["content"]
    # read are dropped: non-dicts, and dicts without string content.
    assert webui_mod._normalize_history([{"role": "user"}]) == []
    assert webui_mod._normalize_history(["hello"]) == []
    assert webui_mod._normalize_history([123]) == []
    assert webui_mod._normalize_history([{"role": "user", "content": None}]) == []
    assert webui_mod._normalize_history("not a list") == []
    # A missing or out-of-set role is coerced to "user" so it can't widen trust.
    assert webui_mod._normalize_history(
        [{"content": "q"}]) == [{"role": "user", "content": "q"}]
    assert webui_mod._normalize_history(
        [{"role": "sneaky", "content": "q"}]) == [{"role": "user", "content": "q"}]
    # Valid roles are preserved and order is kept.
    assert webui_mod._normalize_history(
        [{"role": "assistant", "content": "a"}, {"role": "user", "content": "b"}]
    ) == [{"role": "assistant", "content": "a"}, {"role": "user", "content": "b"}]


def test_chat_rejects_malformed_history(server):
    # A hand-crafted POST with non-dict items or missing content used to reach
    # the backend and raise KeyError/TypeError *after* the chunked stream had
    # started; now it is rejected with a clean 400 before streaming.
    base, _ = server
    for bad in ([{"role": "user"}], ["hello"], [123],
                [{"role": "user", "content": None}]):
        with pytest.raises(urllib.error.HTTPError) as err:
            _post(base + "/api/chat", {"messages": bad})
        assert err.value.code == 400


def test_normalize_answers_coerces_untrusted_input():
    # Well-formed two-element pairs pass through, each side stringified.
    assert webui_mod._normalize_answers(
        [["When due?", "Friday"], ("Dept?", 5)]
    ) == [("When due?", "Friday"), ("Dept?", "5")]
    # Items that would crash "for q, a in answers" are dropped instead:
    # non-iterables, wrong-length sequences, and non-list input.
    assert webui_mod._normalize_answers([5]) == []
    assert webui_mod._normalize_answers(["x"]) == []
    assert webui_mod._normalize_answers([[1, 2, 3]]) == []
    assert webui_mod._normalize_answers([{"q": "a"}]) == []
    assert webui_mod._normalize_answers("not a list") == []
    # Valid pairs survive even when interleaved with junk; order is kept.
    assert webui_mod._normalize_answers(
        [["q1", "a1"], 5, ["q2", "a2"]]) == [("q1", "a1"), ("q2", "a2")]


def test_tender_finish_tolerates_malformed_answers(server):
    # A hand-crafted POST whose answers carry non-pair items used to unpack
    # straight into TypeError/ValueError and surface as an opaque 500; the
    # boundary now drops the junk and still builds a checklist (HTTP 200).
    base, ui = server
    status, result = _post(
        base + "/api/tender/finish",
        {"item": "Firewalls", "answers": [["When due?", "Friday"], 5, ["x"]]})
    assert status == 200
    assert result["count"] == len(CHECKLIST["requirements"])
    assert (ui.out_dir / result["file"]).exists()
