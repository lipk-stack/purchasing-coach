import re
import zipfile

import pytest
from openpyxl import load_workbook

import coach.excel as excel
from coach.excel import (
    BRIEF_SHEET,
    INFO_SHEET,
    REVIEW_DECISION_OPTIONS,
    REVIEW_SHEET,
    VENDOR_STATUS_OPTIONS,
    write_checklist,
)
from coach.models import RequirementRow, TenderInfo

INFO = TenderInfo(
    issue_date="2026-06-10",
    submission_deadline="2026-07-10",
    purchase_item="Endpoint backup SaaS",
    issued_by="IT Procurement",
    requesting_dept="Infrastructure",
    tender_reference="XXEON-IT-2026-014",
    procurement_type="Tender",
    estimated_value="MYR 250,000",
    purchase_category="Cloud Services",
)

ROWS = [
    RequirementRow(ref="5.3", section="Access Control",
                   requirement="Enforce MFA for all user accounts.", mandatory="M"),
    RequirementRow(ref="5.6", section="Audits and Assessments",
                   requirement="Provide annual SOC 2 Type II report.", mandatory="M"),
    RequirementRow(ref="10.1", section="TCO Analysis",
                   requirement="Provide 5-year TCO analysis.", mandatory="O"),
]


def test_write_with_template(samples, tmp_path):
    out = write_checklist(INFO, ROWS, tmp_path / "out.xlsx", samples["template"])
    wb = load_workbook(out)

    info = wb["Tender Information"]
    cells = {info.cell(r, 1).value: info.cell(r, 2).value
             for r in range(1, info.max_row + 1)}
    assert cells["Purchase Item"] == "Endpoint backup SaaS"
    assert cells["Tender Reference"] == "XXEON-IT-2026-014"

    tracker = wb["Compliance Tracker"]
    rows = list(tracker.iter_rows(values_only=True))
    header_idx = next(i for i, r in enumerate(rows) if r and "Seq" in r)
    data = rows[header_idx + 1: header_idx + 1 + len(ROWS)]
    assert data[0][:5] == (1, "5.3", "Access Control",
                           "Enforce MFA for all user accounts.", "M")
    assert data[2][1] == "10.1"


def test_oversize_template_is_refused(monkeypatch, tmp_path):
    # A .xlsx is a zip; a member that decompresses past the cap (tiny on disk,
    # huge expanded) is a zip bomb and must be refused before openpyxl parses it.
    monkeypatch.setattr(excel, "_MAX_TEMPLATE_UNCOMPRESSED_BYTES", 2000)
    bomb = tmp_path / "bomb.xlsx"
    with zipfile.ZipFile(bomb, "w", zipfile.ZIP_DEFLATED) as zf:
        # Highly compressible payload: tiny on disk, far over the shrunk cap.
        zf.writestr("xl/worksheets/sheet1.xml", "A" * 200_000)
    assert bomb.stat().st_size < 2000  # actually tiny on disk
    with pytest.raises(ValueError, match="zip bomb|too large to process"):
        write_checklist(INFO, ROWS, tmp_path / "out.xlsx", bomb)


def test_corrupt_template_is_refused(tmp_path):
    not_a_zip = tmp_path / "broken.xlsx"
    not_a_zip.write_bytes(b"this is not a zip file")
    with pytest.raises(ValueError, match="not a valid .xlsx"):
        write_checklist(INFO, ROWS, tmp_path / "out.xlsx", not_a_zip)


def test_write_without_template(tmp_path):
    out = write_checklist(INFO, ROWS, tmp_path / "no_template.xlsx", None)
    wb = load_workbook(out)
    assert "Tender Information" in wb.sheetnames
    assert "Compliance Tracker" in wb.sheetnames


def _status_column_letter(tracker):
    for row in tracker.iter_rows(max_row=20):
        for cell in row:
            if cell.value and str(cell.value).strip() == "Vendor Status":
                return cell.column_letter, cell.row
    raise AssertionError("Vendor Status header not found")


def test_vendor_status_dropdown_and_freeze(samples, tmp_path):
    out = write_checklist(INFO, ROWS, tmp_path / "dv.xlsx", samples["template"])
    wb = load_workbook(out)
    tracker = wb["Compliance Tracker"]

    letter, header_row = _status_column_letter(tracker)
    # A list validation covers the written Vendor Status cells with the
    # standard compliance vocabulary.
    last_row = header_row + len(ROWS)
    target = f"{letter}{header_row + 1}:{letter}{last_row}"
    dv = next((d for d in tracker.data_validations.dataValidation
               if d.type == "list" and target in str(d.sqref)), None)
    assert dv is not None
    for option in VENDOR_STATUS_OPTIONS:
        assert option in dv.formula1
    # Header stays visible while scrolling the checklist.
    assert tracker.freeze_panes == f"A{header_row + 1}"


def test_status_dropdown_without_template(tmp_path):
    out = write_checklist(INFO, ROWS, tmp_path / "dv_blank.xlsx", None)
    wb = load_workbook(out)
    tracker = wb["Compliance Tracker"]
    dvs = list(tracker.data_validations.dataValidation)
    assert any(d.type == "list" for d in dvs)


def test_review_sheet_summary_and_signoff(samples, tmp_path):
    out = write_checklist(INFO, ROWS, tmp_path / "review.xlsx", samples["template"])
    wb = load_workbook(out)
    assert REVIEW_SHEET in wb.sheetnames
    ws = wb[REVIEW_SHEET]

    labels = {ws.cell(r, 1).value: ws.cell(r, 2).value
              for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value}

    # Live count over the tracker's Vendor Status column (data rows only).
    letter, header_row = _status_column_letter(wb["Compliance Tracker"])
    last_row = header_row + len(ROWS)
    rng = f"'Compliance Tracker'!{letter}{header_row + 1}:{letter}{last_row}"
    assert labels["Compliant"] == f'=COUNTIF({rng},"Compliant")'
    assert labels["Total requirements"] == len(ROWS)
    # The go/no-go figure: mandatory requirements the vendor marked non-compliant.
    assert 'COUNTIFS' in labels["Mandatory non-compliant (review blocker)"]
    assert '"Non-Compliant"' in labels["Mandatory non-compliant (review blocker)"]

    # Reviewer sign-off block with a fixed decision vocabulary.
    assert "Approved By" in labels
    decision_dv = next(d for d in ws.data_validations.dataValidation
                       if d.type == "list")
    for option in REVIEW_DECISION_OPTIONS:
        assert option in decision_dv.formula1


def test_review_sheet_without_template(tmp_path):
    out = write_checklist(INFO, ROWS, tmp_path / "review_blank.xlsx", None)
    wb = load_workbook(out)
    assert REVIEW_SHEET in wb.sheetnames


def test_review_sheet_compliance_rate_and_blocker_formatting(samples, tmp_path):
    out = write_checklist(INFO, ROWS, tmp_path / "rate.xlsx", samples["template"])
    wb = load_workbook(out)
    ws = wb[REVIEW_SHEET]

    rows = {ws.cell(r, 1).value: r for r in range(1, ws.max_row + 1)
            if ws.cell(r, 1).value}

    # A live compliance-rate row, shown as a percentage, that divides Compliant
    # by the applicable (non-N/A) rows and is divide-by-zero safe.
    rate_row = rows["Compliance rate (of applicable)"]
    rate_cell = ws.cell(rate_row, 2)
    assert str(rate_cell.value).startswith("=IFERROR(COUNTIF(")
    assert '"Not Applicable"' in rate_cell.value
    assert rate_cell.number_format == "0.0%"

    # The rate cell carries a green data-bar gauge fixed to a 0%..100% scale so
    # the bar length means the same on every workbook.
    rate_target = f"B{rate_row}"
    data_bars = [r for rng, rules in ws.conditional_formatting._cf_rules.items()
                 for r in rules
                 if r.type == "dataBar" and rate_target in str(rng.sqref)]
    assert len(data_bars) == 1
    cfvo = data_bars[0].dataBar.cfvo
    assert [(c.type, c.val) for c in cfvo] == [("num", 0.0), ("num", 1.0)]

    # The mandatory non-compliant cell is flagged by conditional formatting:
    # red when > 0 (a review blocker), green at 0.
    blocker_row = rows["Mandatory non-compliant (review blocker)"]
    target = f"B{blocker_row}"
    rules = [r for rng, rules in ws.conditional_formatting._cf_rules.items()
             for r in rules if target in str(rng.sqref)]
    operators = {r.operator for r in rules}
    assert "greaterThan" in operators and "equal" in operators


# --- Live-formula evaluation -------------------------------------------------
# The tests above check the formula *strings*; openpyxl never evaluates them and
# the headless-LibreOffice path is unavailable in CI, so a wrong range or an
# off-by-one in the COUNTIF/COUNTIFS construction would pass silently. The
# helpers below evaluate the bounded set of formula shapes the Review sheet
# actually emits (COUNTIF / COUNTIFS / COUNTBLANK / IFERROR-division) against the
# real tracker cells, so the summary's *computed* numbers are asserted, not just
# its text.

_RANGE = re.compile(r"'([^']+)'!([A-Z]+)(\d+):([A-Z]+)(\d+)")


def _resolve_range(wb, ref):
    sheet, c1, r1, _c2, r2 = _RANGE.fullmatch(ref).groups()
    ws = wb[sheet]
    return [ws[f"{c1}{r}"].value for r in range(int(r1), int(r2) + 1)]


def _eval_count(wb, call):
    name, args = call
    if name == "COUNTBLANK":
        return sum(1 for v in _resolve_range(wb, args[0])
                   if v is None or v == "")
    if name == "COUNTIF":
        target = args[1].strip('"')
        return sum(1 for v in _resolve_range(wb, args[0]) if v == target)
    if name == "COUNTIFS":
        rng1, crit1, rng2, crit2 = args
        c1, c2 = crit1.strip('"'), crit2.strip('"')
        return sum(1 for a, b in zip(_resolve_range(wb, rng1),
                                     _resolve_range(wb, rng2), strict=True)
                   if a == c1 and b == c2)
    raise AssertionError(f"unsupported function {name}")


_CALL = re.compile(r"(COUNTIFS|COUNTIF|COUNTBLANK)\(([^()]*)\)")


def eval_formula(wb, formula):
    """Evaluate a Review-sheet summary formula against the live workbook."""
    expr = str(formula).lstrip("=")
    # Reduce every COUNT* call to its integer result (COUNTIFS before COUNTIF).
    while True:
        m = _CALL.search(expr)
        if not m:
            break
        args = [a.strip() for a in m.group(2).split(",")]
        # Re-pair (range, "criterion") tokens since criteria contain no commas.
        expr = expr[:m.start()] + str(_eval_count(wb, (m.group(1), args))) \
            + expr[m.end():]
    ie = re.fullmatch(r"IFERROR\((.*),(\d+)\)", expr)
    if ie:
        try:
            return eval(ie.group(1))  # noqa: S307 - bounded numeric expr
        except ZeroDivisionError:
            return int(ie.group(2))
    return eval(expr)  # noqa: S307 - bounded numeric expr


def test_review_formulas_compute_correct_values(samples, tmp_path):
    rows = [
        RequirementRow(ref="4.1", section="Contract", requirement="a", mandatory="M"),
        RequirementRow(ref="5.1", section="Security", requirement="b", mandatory="M"),
        RequirementRow(ref="5.3", section="Access", requirement="c", mandatory="M"),
        RequirementRow(ref="7.1", section="Support", requirement="d", mandatory="O"),
        RequirementRow(ref="10.1", section="TCO", requirement="e", mandatory="O"),
        RequirementRow(ref="11.1", section="Compliance", requirement="f", mandatory="M"),
    ]
    out = write_checklist(INFO, rows, tmp_path / "eval.xlsx", samples["template"])
    wb = load_workbook(out)
    tracker = wb["Compliance Tracker"]
    status_letter, header_row = _status_column_letter(tracker)

    # Fill a known distribution into the Vendor Status column:
    #   mandatory 5.1 -> Non-Compliant (a review blocker), 10.1 -> Not Applicable,
    #   7.1 left blank (awaiting), the remaining three Compliant.
    status_by_ref = {
        "4.1": "Compliant", "5.1": "Non-Compliant", "5.3": "Compliant",
        "7.1": None, "10.1": "Not Applicable", "11.1": "Compliant",
    }
    ref_col = next(c.column for c in tracker[header_row]
                   if c.value and str(c.value).strip() == "Ref")
    for r in range(header_row + 1, header_row + 1 + len(rows)):
        ref = str(tracker.cell(r, ref_col).value)
        tracker[f"{status_letter}{r}"] = status_by_ref[ref]

    ws = wb[REVIEW_SHEET]
    summary = {ws.cell(r, 1).value: ws.cell(r, 2).value
               for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value}

    assert eval_formula(wb, summary["Compliant"]) == 3
    assert eval_formula(wb, summary["Non-Compliant"]) == 1
    assert eval_formula(wb, summary["Not Applicable"]) == 1
    assert eval_formula(wb, summary["Awaiting vendor response"]) == 1
    assert eval_formula(
        wb, summary["Mandatory non-compliant (review blocker)"]) == 1
    # Compliance rate = Compliant / (total - Not Applicable) = 3 / (6 - 1).
    assert eval_formula(
        wb, summary["Compliance rate (of applicable)"]) == 3 / 5


def test_review_compliance_rate_is_divide_by_zero_safe(samples, tmp_path):
    out = write_checklist(INFO, ROWS, tmp_path / "empty.xlsx", samples["template"])
    wb = load_workbook(out)
    # Vendor hasn't responded: all statuses blank. Marking every row Not
    # Applicable drives the rate denominator to zero; IFERROR must yield 0.
    tracker = wb["Compliance Tracker"]
    status_letter, header_row = _status_column_letter(tracker)
    for r in range(header_row + 1, header_row + 1 + len(ROWS)):
        tracker[f"{status_letter}{r}"] = "Not Applicable"
    ws = wb[REVIEW_SHEET]
    summary = {ws.cell(r, 1).value: ws.cell(r, 2).value
               for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value}
    assert eval_formula(wb, summary["Compliance rate (of applicable)"]) == 0


INTERVIEW = [
    ("Will the software handle personal or payment data?", "Yes - personal data"),
    ("What payment schedule is preferred?", "Quarterly"),
    ("Does it include physical hardware?", "No hardware"),
]


def test_brief_sheet_records_interview(samples, tmp_path):
    out = write_checklist(INFO, ROWS, tmp_path / "brief.xlsx",
                          samples["template"], interview=INTERVIEW)
    wb = load_workbook(out)
    assert BRIEF_SHEET in wb.sheetnames
    # Sits immediately after the Tender Information sheet for logical reading.
    names = wb.sheetnames
    assert names.index(BRIEF_SHEET) == names.index(INFO_SHEET) + 1

    ws = wb[BRIEF_SHEET]
    flat = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
    blob = " | ".join(flat)
    # Purchase context is carried over from the tender info.
    assert "Endpoint backup SaaS" in blob
    assert INFO.purchase_category in blob
    # Every interview question and answer is captured verbatim.
    for question, answer in INTERVIEW:
        assert question in blob
        assert answer in blob

    # The Q&A is laid out as numbered rows under a header.
    header = next(r for r in ws.iter_rows(values_only=True)
                  if r and r[0] == "#")
    assert header[1] == "Question" and header[2] == "Response"


def test_brief_sheet_omitted_without_interview(samples, tmp_path):
    # Existing callers that don't pass an interview get the original workbook.
    out = write_checklist(INFO, ROWS, tmp_path / "no_brief.xlsx",
                          samples["template"])
    wb = load_workbook(out)
    assert BRIEF_SHEET not in wb.sheetnames


def test_brief_sheet_idempotent_on_rerun(samples, tmp_path):
    # Writing twice to the same path must not stack duplicate Brief sheets.
    path = tmp_path / "rerun.xlsx"
    write_checklist(INFO, ROWS, path, samples["template"], interview=INTERVIEW)
    write_checklist(INFO, ROWS, path, path, interview=INTERVIEW)
    wb = load_workbook(path)
    assert wb.sheetnames.count(BRIEF_SHEET) == 1
