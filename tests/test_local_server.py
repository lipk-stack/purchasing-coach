"""Integration test: full flow over real HTTP against a mock LM Studio server."""

import json
import threading
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer

import pytest
from openpyxl import load_workbook

from coach.backends import OpenAICompatBackend
from coach.backends import openai_compat
from coach.backends.openai_compat import BackendError
from coach.llm import Coach
from coach.tender import run_tender_flow

PLAN = {"questions": [{"key": "deadline",
                       "question": "When is the submission deadline?"}]}
CHECKLIST = {
    "tender_info": {"purchase_item": "Backup SaaS",
                    "tender_reference": "REF-1"},
    "requirements": [{"ref": "5.6", "section": "Audits",
                      "requirement": "Provide annual SOC 2 Type II report.",
                      "mandatory": "M"}],
}


class MockLMStudio(BaseHTTPRequestHandler):
    def log_message(self, *args):  # keep test output quiet
        pass

    def do_GET(self):
        # The backend probes LM Studio's native endpoint first; this mock only
        # speaks the OpenAI-compatible API, so it 404s there and the backend
        # falls back to /v1/models — exactly the Ollama/plain-server path.
        if self.path == "/api/v0/models":
            self.send_error(404)
            return
        assert self.path == "/v1/models"
        self._json({"data": [{"id": "test-model-7b"}]})

    def do_POST(self):
        assert self.path == "/v1/chat/completions"
        length = int(self.headers["Content-Length"])
        payload = json.loads(self.rfile.read(length))
        if payload.get("stream"):
            self._sse(["Per clause 5.6, ", "SOC 2 is required."])
            return
        # Structured request: decide by schema name sent in response_format.
        schema_name = payload["response_format"]["json_schema"]["name"]
        content = PLAN if schema_name == "interview_plan" else CHECKLIST
        self._json({"choices": [{"message": {
            "content": json.dumps(content)}}]})

    def _json(self, body: dict):
        data = json.dumps(body).encode()
        self.send_response(200)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def _sse(self, parts: list[str]):
        self.send_response(200)
        self.send_header("Content-Type", "text/event-stream")
        self.end_headers()
        for part in parts:
            chunk = {"choices": [{"delta": {"content": part}}]}
            self.wfile.write(f"data: {json.dumps(chunk)}\n\n".encode())
        self.wfile.write(b"data: [DONE]\n\n")


@pytest.fixture()
def server():
    httpd = ThreadingHTTPServer(("127.0.0.1", 0), MockLMStudio)
    thread = threading.Thread(target=httpd.serve_forever, daemon=True)
    thread.start()
    yield f"http://127.0.0.1:{httpd.server_address[1]}/v1"
    httpd.shutdown()


def test_full_flow_over_http(server, tmp_path):
    backend = OpenAICompatBackend(server, name="lmstudio")
    assert backend.model == "test-model-7b"  # picked up from /models

    coach = Coach("guideline text", backend)

    reply = "".join(coach.answer([{"role": "user", "content": "SOC 2?"}]))
    assert reply == "Per clause 5.6, SOC 2 is required."

    scripted = iter(["Backup SaaS for endpoints", "1 July 2026"])
    out = run_tender_flow(coach, template_path=None, out_dir=tmp_path,
                          ask=lambda prompt: next(scripted),
                          say=lambda *a: None)
    wb = load_workbook(out)
    flat = [c for row in wb["Compliance Tracker"].iter_rows(values_only=True)
            for c in row if c]
    assert "Provide annual SOC 2 Type II report." in flat
    info = {wb["Tender Information"].cell(r, 1).value:
            wb["Tender Information"].cell(r, 2).value
            for r in range(1, wb["Tender Information"].max_row + 1)}
    assert info["Purchase Item"] == "Backup SaaS"
    assert info["Issue Date"] == "TBC"  # unanswered fields default to TBC


class FloodServer(BaseHTTPRequestHandler):
    """A hostile server that answers /v1/models with an over-long body."""

    def log_message(self, *args):
        pass

    def do_GET(self):
        if self.path == "/api/v0/models":
            self.send_error(404)
            return
        # Far larger than the (test-shrunk) cap, to prove the read is bounded.
        body = b'{"data": [' + b'{"id": "x"},' * 100_000 + b'{"id": "y"}]}'
        self.send_response(200)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


@pytest.fixture()
def flood_server():
    httpd = ThreadingHTTPServer(("127.0.0.1", 0), FloodServer)
    thread = threading.Thread(target=httpd.serve_forever, daemon=True)
    thread.start()
    yield f"http://127.0.0.1:{httpd.server_address[1]}/v1"
    httpd.shutdown()


def test_oversize_response_is_refused(flood_server, monkeypatch):
    # Shrink the cap so the test stays small and fast; the guard is the same.
    monkeypatch.setattr(openai_compat, "MAX_RESPONSE_BYTES", 4096)
    be = OpenAICompatBackend(flood_server, model="m")  # skip model discovery
    with pytest.raises(BackendError, match="implausibly large"):
        be.list_models()
