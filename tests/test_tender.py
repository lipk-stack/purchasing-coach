"""End-to-end tender flow with a fake LLM backend (no network)."""

from openpyxl import load_workbook

from coach.llm import Coach
from coach.tender import run_tender_flow

PLAN = {
    "questions": [
        {"key": "deadline", "question": "When is the submission deadline?"},
        {"key": "dept", "question": "Which department is requesting this?"},
    ]
}

CHECKLIST = {
    "tender_info": {
        "issue_date": "2026-06-10", "submission_deadline": "2026-07-01",
        "purchase_item": "Firewall appliances", "issued_by": "IT Procurement",
        "requesting_dept": "Network", "tender_reference": "XXEON-IT-2026-001",
        "procurement_type": "Tender", "estimated_value": "MYR 400,000",
        "purchase_category": "Hardware",
    },
    "requirements": [
        # Out of order, with a paraphrased section and a hallucinated clause,
        # so the reconciliation path is exercised end to end.
        {"ref": "8.4", "section": "Warranty (paraphrased)",
         "requirement": "Declare End-of-Sale and End-of-Support dates.",
         "mandatory": "M"},
        {"ref": "5.6", "section": "Audits",
         "requirement": "Supply an annual SOC 2 Type II report.",
         "mandatory": "M"},
        {"ref": "99.9", "section": "Imaginary",
         "requirement": "A requirement that is not in the guideline.",
         "mandatory": "O"},
    ],
}

GUIDELINE = """\
## 5 INFORMATION SECURITY CONSIDERATIONS

### 5.6 Audits and Assessments

## 8 HARDWARE REQUIREMENTS

### 8.4 Warranty and Replacement Policies
"""


class FakeBackend:
    name = "fake"
    model = "fake-model"

    def stream_chat(self, system, messages, max_tokens=4096):
        assert "guideline" in system
        yield from ["See clause 8.4 ", "(Warranty)."]

    def complete_json(self, system, prompt, schema, schema_name,
                      max_tokens=8192):
        return PLAN if schema_name == "interview_plan" else CHECKLIST


def make_coach():
    return Coach(GUIDELINE, FakeBackend())


def test_chat_answer_streams():
    coach = make_coach()
    reply = "".join(coach.answer([{"role": "user", "content": "warranty?"}]))
    assert reply == "See clause 8.4 (Warranty)."


def test_tender_flow_writes_workbook(tmp_path):
    coach = make_coach()
    # The interview now merges guideline-coverage questions, so answer every
    # prompt rather than scripting a fixed count.
    out = run_tender_flow(coach, template_path=None, out_dir=tmp_path,
                          ask=lambda prompt: "Firewall appliances",
                          say=lambda *a: None)
    assert out is not None and out.exists()

    wb = load_workbook(out)
    tracker = wb["Compliance Tracker"]
    flat = [c for row in tracker.iter_rows(values_only=True) for c in row if c]
    assert "Declare End-of-Sale and End-of-Support dates." in flat
    # Section titles are canonicalised to the real guideline headings.
    assert "Warranty and Replacement Policies" in flat
    assert "Audits and Assessments" in flat


def test_tender_flow_reconciles_against_guideline(tmp_path):
    coach = make_coach()
    notes: list[str] = []
    out = run_tender_flow(coach, None, tmp_path,
                          ask=lambda prompt: "Firewall appliances",
                          say=lambda *a: notes.append(" ".join(map(str, a))))
    wb = load_workbook(out)
    tracker = wb["Compliance Tracker"]
    refs = [row[1] for row in tracker.iter_rows(min_row=4, values_only=True)
            if row[1]]
    # Rows are reordered into guideline order (5.6 before 8.4 before 99.9).
    assert refs == ["5.6", "8.4", "99.9"]
    # The hallucinated clause is flagged for the user.
    assert any("99.9" in n and "could not be matched" in n for n in notes)


GRANULAR_GUIDELINE = """\
## 5 INFORMATION SECURITY CONSIDERATIONS

### 5.3 Access Control and Authentication Requirements

Multi-factor authentication (MFA) must be enforced for all user accounts.

Role-based access control (RBAC) must be implemented.

Integration with XXEON Single Sign-On (SSO) is required where applicable.

### 5.6 Audits and Assessments

Annual third-party security audits are mandatory.

Right-to-audit clauses must be included in all vendor agreements.
"""


class GranularBackend(FakeBackend):
    # The model selects whole section 5; expansion fans it out to every
    # sub-clause requirement parsed from the guideline body.
    def complete_json(self, system, prompt, schema, schema_name,
                      max_tokens=8192):
        if schema_name == "interview_plan":
            return PLAN
        return {
            "tender_info": CHECKLIST["tender_info"],
            "requirements": [
                {"ref": "5", "section": "Information Security",
                 "requirement": "Information security applies.",
                 "mandatory": "M"},
            ],
        }


def test_checklist_expands_section_into_granular_rows():
    coach = Coach(GRANULAR_GUIDELINE, GranularBackend())
    checklist = coach.build_checklist("SaaS analytics tool", [("q", "a")])
    reqs = checklist.requirements
    # Citing section "5" fans out to all five sub-clause requirements, verbatim
    # from the guideline body and in guideline order.
    statements = [r.requirement for r in reqs]
    assert "Multi-factor authentication (MFA) must be enforced for all " \
           "user accounts." in statements
    assert "Right-to-audit clauses must be included in all vendor " \
           "agreements." in statements
    assert len(reqs) == 5
    # Sections are the real clause titles; rows ordered 5.3 before 5.6.
    assert [r.ref for r in reqs] == ["5.3", "5.3", "5.3", "5.6", "5.6"]
    assert reqs[0].section == "Access Control and Authentication Requirements"
    assert all(r.mandatory in ("M", "O") for r in reqs)


def test_interview_adds_guideline_coverage_questions():
    coach = Coach(GRANULAR_GUIDELINE, GranularBackend())
    plan = coach.plan_interview("SaaS analytics tool")
    text = " ".join(q.question.lower() for q in plan.questions)
    # The model's two questions plus merged coverage for the data/security
    # section that the guideline actually contains.
    assert len(plan.questions) > len(PLAN["questions"])
    assert "personal data" in text


SAFETYNET_GUIDELINE = """\
## 4 CONTRACT REQUIREMENTS

### 4.1 Standard Contract Terms

The vendor must define all deliverables and acceptance criteria.

## 5 INFORMATION SECURITY CONSIDERATIONS

### 5.6 Audits and Assessments

Annual third-party security audits are mandatory.

## 8 HARDWARE REQUIREMENTS

### 8.4 Warranty and Replacement Policies

Minimum warranty periods must be specified.

## 11 COMPLIANCE AND RISK MANAGEMENT

### 11.1 Regulatory Compliance

The vendor must comply with all applicable regulations.
"""


class UnderSelectingBackend(FakeBackend):
    # A small model that only picks the obvious hardware clause and drops the
    # cross-cutting compliance sections — the safety net must add them back.
    def complete_json(self, system, prompt, schema, schema_name,
                      max_tokens=8192):
        if schema_name == "interview_plan":
            return PLAN
        return {
            "tender_info": CHECKLIST["tender_info"],
            "requirements": [
                {"ref": "8.4", "section": "Warranty",
                 "requirement": "Specify warranty periods.", "mandatory": "M"},
            ],
        }


def test_safety_net_adds_core_sections_when_model_under_selects(tmp_path):
    coach = Coach(SAFETYNET_GUIDELINE, UnderSelectingBackend())
    checklist = coach.build_checklist("Firewall appliances", [("q", "a")])
    refs = [r.ref for r in checklist.requirements]
    # Core sections 4, 5 and 11 are folded in despite the model only naming 8.4.
    assert refs == ["4.1", "5.6", "8.4", "11.1"]
    assert checklist.added_core_sections == ["4", "5", "11"]


def test_safety_net_note_surfaced_in_tender_flow(tmp_path):
    coach = Coach(SAFETYNET_GUIDELINE, UnderSelectingBackend())
    notes: list[str] = []
    run_tender_flow(coach, None, tmp_path,
                    ask=lambda prompt: "Firewall appliances",
                    say=lambda *a: notes.append(" ".join(map(str, a))))
    assert any("added automatically" in n and "4, 5, 11" in n for n in notes)


class CoreOnlyBackend(FakeBackend):
    # A model that selects nothing item-specific — only a core security clause.
    # The buyer's interview answers must still pull in the hardware section.
    def complete_json(self, system, prompt, schema, schema_name,
                      max_tokens=8192):
        if schema_name == "interview_plan":
            return PLAN
        return {
            "tender_info": CHECKLIST["tender_info"],
            "requirements": [
                {"ref": "5.6", "section": "Audits",
                 "requirement": "Annual audits.", "mandatory": "M"},
            ],
        }


def test_interview_answers_drive_section_inclusion():
    coach = Coach(SAFETYNET_GUIDELINE, CoreOnlyBackend())
    answers = [
        ("Does this purchase include physical hardware or equipment?",
         "Yes, firewall appliances"),
    ]
    checklist = coach.build_checklist("Firewall appliances", answers)
    refs = [r.ref for r in checklist.requirements]
    # Section 8 is folded in because the buyer affirmed hardware, even though
    # the model never selected it; core 4/11 come in too.
    assert refs == ["4.1", "5.6", "8.4", "11.1"]
    assert checklist.added_core_sections == ["4", "8", "11"]


def test_negative_answer_does_not_add_item_section():
    coach = Coach(SAFETYNET_GUIDELINE, CoreOnlyBackend())
    answers = [
        ("Does this purchase include physical hardware or equipment?",
         "No, this is a pure SaaS subscription"),
    ]
    checklist = coach.build_checklist("SaaS tool", answers)
    refs = [r.ref for r in checklist.requirements]
    # Hardware (8) is pruned by the negative answer; only core sections added.
    assert "8.4" not in refs
    assert checklist.added_core_sections == ["4", "11"]


def test_tender_flow_cancels_on_empty_item(tmp_path):
    coach = make_coach()
    out = run_tender_flow(coach, None, tmp_path,
                          ask=lambda prompt: "", say=lambda *a: None)
    assert out is None
