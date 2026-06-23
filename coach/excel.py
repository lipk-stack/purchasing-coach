"""Read/write the tender checklist Excel workbook based on the template."""

import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .models import RequirementRow, TenderInfo

# A .xlsx is a zip (Office Open XML), so the user-supplied ``--template`` is a
# compressed input with the same decompression-amplification / zip-bomb risk the
# .docx guideline loader guards against (see coach/documents.py). A file tiny on
# disk but enormous when expanded could exhaust memory inside ``load_workbook``
# before any of our code runs. We validate the archive's total decompressed size
# with a bounded read first, so such a file is rejected with a clear error rather
# than OOMing the machine. The real TENDER_TEMPLATE.xlsx is ~18 KB; 128 MiB is
# vast headroom for any legitimate template while still bounding the damage.
_MAX_TEMPLATE_UNCOMPRESSED_BYTES = 128 * 1024 * 1024
_TEMPLATE_READ_CHUNK = 1024 * 1024

INFO_SHEET = "Tender Information"
TRACKER_SHEET = "Compliance Tracker"
REVIEW_SHEET = "Review & Approval"
BRIEF_SHEET = "Procurement Brief"
TRACKER_HEADERS = ["Seq", "Ref", "Section", "Requirement", "M/O",
                   "Vendor Status", "Vendor Remarks"]

# The reviewer's recommendation is also a fixed vocabulary so the approval
# decision is unambiguous and filterable across submissions.
REVIEW_DECISION_OPTIONS = [
    "Approved",
    "Approved with Conditions",
    "Rejected",
    "Resubmission Required",
]

# Standardised values the vendor/service provider picks from when populating
# the "Vendor Status" column. A fixed vocabulary (rather than free text) keeps
# the submitted checklist consistent and lets the reviewer filter/score it
# during review and approval. Free-text justification still goes in the
# adjacent "Vendor Remarks" column.
VENDOR_STATUS_OPTIONS = [
    "Compliant",
    "Partially Compliant",
    "Non-Compliant",
    "Not Applicable",
]


def write_checklist(
    tender_info: TenderInfo,
    requirements: list[RequirementRow],
    out_path: str | Path,
    template_path: str | Path | None = None,
    interview: list[tuple[str, str]] | None = None,
) -> Path:
    """Fill the tender template with the interview results and save it.

    If ``template_path`` is given the workbook is loaded from it (preserving
    its formatting); otherwise a fresh workbook with the same layout is built.

    ``interview`` is the reverse-prompting question/answer record. When given,
    it is captured on a ``Procurement Brief`` sheet so the reviewer/approver can
    see the buyer's declared requirements — the basis on which the compliance
    scope was selected.
    """
    out_path = Path(out_path)
    if template_path and Path(template_path).exists():
        wb = _load_template_workbook(template_path)
    else:
        wb = create_blank_template()

    _fill_info_sheet(wb, tender_info)
    header_row, last_row, col = _fill_tracker_sheet(wb, requirements)
    _add_review_sheet(wb, header_row, last_row, col)
    _add_brief_sheet(wb, tender_info, interview)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path


def _fill_info_sheet(wb, tender_info: TenderInfo) -> None:
    ws = _find_sheet(wb, INFO_SHEET)
    values = {label: getattr(tender_info, attr)
              for label, attr in TenderInfo.TEMPLATE_LABELS.items()}
    found = set()
    for row in ws.iter_rows():
        for cell in row:
            label = str(cell.value).strip() if cell.value else ""
            if label in values:
                target = ws.cell(row=cell.row, column=cell.column + 1)
                target.value = values[label]
                target.alignment = Alignment(vertical="top", wrap_text=True)
                found.add(label)
    # Labels missing from the template are appended at the bottom.
    for label in values:
        if label not in found:
            row = ws.max_row + 1
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=values[label])


def _fill_tracker_sheet(wb, requirements: list[RequirementRow]):
    ws = _find_sheet(wb, TRACKER_SHEET)
    header_row = _find_header_row(ws)
    col = {str(c.value).strip(): c.column for c in ws[header_row] if c.value}

    row = header_row + 1
    for seq, req in enumerate(requirements, start=1):
        entries = {
            "Seq": seq,
            "Ref": req.ref,
            "Section": req.section,
            "Requirement": req.requirement,
            "M/O": req.mandatory,
            "Vendor Status": "",
            "Vendor Remarks": "",
        }
        for header, value in entries.items():
            if header in col:
                cell = ws.cell(row=row, column=col[header], value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        row += 1

    last_row = row - 1
    _extend_table(ws, header_row, last_row)
    _add_status_dropdown(ws, header_row, last_row, col)
    # Keep the header (and the tender title above it) visible while the
    # reviewer scrolls a long, granular checklist.
    if last_row > header_row:
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    return header_row, last_row, col


def _add_status_dropdown(ws, header_row: int, last_row: int,
                         col: dict[str, int]) -> None:
    """Constrain the Vendor Status column to the standard compliance values.

    The vendor populates this column with a dropdown choice, so submissions are
    consistent and the reviewer can filter and approve against a fixed
    vocabulary. No-op when the column is absent or there are no data rows.
    """
    status_col = col.get("Vendor Status")
    if not status_col or last_row <= header_row:
        return
    letter = get_column_letter(status_col)
    options = ",".join(VENDOR_STATUS_OPTIONS)
    dv = DataValidation(type="list", formula1=f'"{options}"',
                        allow_blank=True, showDropDown=False)
    dv.error = "Pick one of the standard compliance statuses."
    dv.errorTitle = "Invalid status"
    dv.prompt = "Select the vendor's compliance status for this requirement."
    dv.promptTitle = "Vendor Status"
    dv.add(f"{letter}{header_row + 1}:{letter}{last_row}")
    ws.add_data_validation(dv)


def _add_review_sheet(wb, header_row: int, last_row: int,
                      col: dict[str, int]) -> None:
    """Add a Review & Approval sheet that tallies the vendor's submission.

    The compliance summary uses live formulas over the Compliance Tracker, so
    the moment the vendor populates the Vendor Status dropdown the reviewer sees
    the counts — including the **mandatory non-compliant** total, the go/no-go
    figure. A reviewer sign-off block (with a fixed decision dropdown) captures
    the approval outcome on the same workbook that's submitted for review.

    No-op when there are no data rows or the status column is absent.
    """
    status_col = col.get("Vendor Status")
    mo_col = col.get("M/O")
    if not status_col or last_row <= header_row:
        return

    tracker = _find_sheet(wb, TRACKER_SHEET)
    # Quote the sheet name for cross-sheet formula references (it has a space).
    sheet = f"'{tracker.title}'"
    status = get_column_letter(status_col)
    srange = f"{sheet}!{status}{header_row + 1}:{status}{last_row}"
    total = last_row - header_row

    def countif(value: str) -> str:
        return f'=COUNTIF({srange},"{value}")'

    summary = [
        ("Total requirements", total),
        ("Compliant", countif("Compliant")),
        ("Partially Compliant", countif("Partially Compliant")),
        ("Non-Compliant", countif("Non-Compliant")),
        ("Not Applicable", countif("Not Applicable")),
        ("Awaiting vendor response", f"=COUNTBLANK({srange})"),
    ]
    if mo_col:
        mo = get_column_letter(mo_col)
        mrange = f"{sheet}!{mo}{header_row + 1}:{mo}{last_row}"
        summary += [
            ("Mandatory (M) requirements", f'=COUNTIF({mrange},"M")'),
            ("Mandatory non-compliant (review blocker)",
             f'=COUNTIFS({mrange},"M",{srange},"Non-Compliant")'),
        ]

    # Live compliance rate: Compliant rows over the applicable ones (excludes
    # Not Applicable so an N/A-heavy submission isn't penalised). Guarded
    # against divide-by-zero while the vendor hasn't filled anything in.
    rate_label = "Compliance rate (of applicable)"
    summary.append(
        (rate_label,
         f'=IFERROR(COUNTIF({srange},"Compliant")'
         f'/({total}-COUNTIF({srange},"Not Applicable")),0)'))
    blocker_label = "Mandatory non-compliant (review blocker)"

    # Build a fresh sheet so re-runs stay idempotent.
    if REVIEW_SHEET in wb.sheetnames:
        del wb[REVIEW_SHEET]
    ws = wb.create_sheet(REVIEW_SHEET)
    title_font = Font(bold=True, size=14)
    label_font = Font(bold=True)
    blank_fill = PatternFill("solid", fgColor="FFF2CC")

    ws["A1"] = "REVIEW & APPROVAL"
    ws["A1"].font = title_font

    ws["A3"] = "Compliance Summary"
    ws["A3"].font = label_font
    ws["B3"] = "(updates live as the vendor fills in Vendor Status)"
    row = 4
    blocker_row = None
    rate_row = None
    for label, value in summary:
        ws.cell(row=row, column=1, value=label).font = label_font
        cell = ws.cell(row=row, column=2, value=value)
        if label == rate_label:
            cell.number_format = "0.0%"
            rate_row = row
        if label == blocker_label:
            blocker_row = row
        row += 1

    # Give the compliance rate an at-a-glance gauge: a green data bar that fills
    # the cell proportionally from 0% to 100% (fixed scale, so the bar means the
    # same on every workbook). Updates live with the underlying IFERROR rate.
    if rate_row is not None:
        ws.conditional_formatting.add(
            f"B{rate_row}",
            DataBarRule(start_type="num", start_value=0,
                        end_type="num", end_value=1, color="63BE7B"))

    # Make the go/no-go figure unmissable: red when any mandatory requirement
    # is non-compliant, green once it clears to zero. Updates live with the
    # underlying COUNTIFS as the vendor fills the tracker.
    if blocker_row is not None:
        target = f"B{blocker_row}"
        red = PatternFill("solid", fgColor="FFC7CE")
        green = PatternFill("solid", fgColor="C6EFCE")
        ws.conditional_formatting.add(
            target, CellIsRule(operator="greaterThan", formula=["0"],
                               fill=red, font=Font(color="9C0006", bold=True)))
        ws.conditional_formatting.add(
            target, CellIsRule(operator="equal", formula=["0"],
                               fill=green, font=Font(color="006100")))

    row += 1
    ws.cell(row=row, column=1, value="Reviewer Sign-off").font = label_font
    row += 1
    signoff = ["Reviewed By", "Review Date", "Approval Decision",
               "Approved By", "Approval Date", "Comments / Conditions"]
    decision_row = None
    for label in signoff:
        ws.cell(row=row, column=1, value=label).font = label_font
        cell = ws.cell(row=row, column=2)
        cell.fill = blank_fill
        if label == "Approval Decision":
            decision_row = row
        row += 1

    if decision_row is not None:
        options = ",".join(REVIEW_DECISION_OPTIONS)
        dv = DataValidation(type="list", formula1=f'"{options}"',
                            allow_blank=True, showDropDown=False)
        dv.promptTitle = "Approval Decision"
        dv.prompt = "Select the review outcome for this submission."
        dv.errorTitle = "Invalid decision"
        dv.error = "Pick one of the standard approval decisions."
        dv.add(f"B{decision_row}")
        ws.add_data_validation(dv)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 40


def _add_brief_sheet(
    wb, tender_info: TenderInfo, interview: list[tuple[str, str]] | None
) -> None:
    """Record the reverse-prompting interview on a Procurement Brief sheet.

    The compliance checklist's *scope* (which guideline sections were included)
    is driven by the buyer's answers during the tender interview. Capturing
    those questions and answers on the same workbook gives the reviewer and
    approver the rationale for the scope — they can confirm the right sections
    were pulled in and challenge any answer that looks wrong — without re-running
    the interview. No-op when there is no interview to record (e.g. a workbook
    written outside the tender flow), so existing callers are unaffected.
    """
    if not interview:
        return

    # Build a fresh sheet so re-runs stay idempotent.
    if BRIEF_SHEET in wb.sheetnames:
        del wb[BRIEF_SHEET]
    ws = wb.create_sheet(BRIEF_SHEET)
    # Read first: sit it right after the Tender Information sheet.
    if INFO_SHEET in wb.sheetnames:
        info_idx = wb.sheetnames.index(INFO_SHEET)
        wb.move_sheet(BRIEF_SHEET, info_idx + 1 - wb.sheetnames.index(BRIEF_SHEET))

    title_font = Font(bold=True, size=14)
    label_font = Font(bold=True)
    wrap_top = Alignment(vertical="top", wrap_text=True)

    ws["A1"] = "PROCUREMENT BRIEF"
    ws["A1"].font = title_font

    ws["A2"] = (
        "The buyer's answers below scoped this compliance checklist. Review "
        "them alongside the Compliance Tracker before approving."
    )
    ws["A2"].font = Font(italic=True, size=9)

    row = 4
    for label, value in (
        ("Purchase Item", tender_info.purchase_item),
        ("Purchase Category", tender_info.purchase_category),
    ):
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=value).alignment = wrap_top
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Interview Record").font = label_font
    row += 1
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    headers = ["#", "Question", "Response"]
    for c, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.font = label_font
        cell.fill = header_fill
    row += 1

    for i, (question, answer) in enumerate(interview, start=1):
        ws.cell(row=row, column=1, value=i).alignment = wrap_top
        ws.cell(row=row, column=2, value=str(question)).alignment = wrap_top
        ws.cell(row=row, column=3, value=str(answer)).alignment = wrap_top
        row += 1

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 45


def _find_sheet(wb, name: str):
    for ws in wb.worksheets:
        if ws.title.strip().lower() == name.lower():
            return ws
    ws = wb.create_sheet(name)
    if name == TRACKER_SHEET:
        for i, header in enumerate(TRACKER_HEADERS, start=1):
            ws.cell(row=1, column=i, value=header).font = Font(bold=True)
    return ws


def _find_header_row(ws) -> int:
    for row in ws.iter_rows(max_row=20):
        values = {str(c.value).strip() for c in row if c.value}
        if "Seq" in values and "Requirement" in values:
            return row[0].row
    # No header found — write one on the first empty row.
    row = ws.max_row + 1 if ws.max_row > 1 or ws["A1"].value else 1
    for i, header in enumerate(TRACKER_HEADERS, start=1):
        ws.cell(row=row, column=i, value=header).font = Font(bold=True)
    return row


def _extend_table(ws, header_row: int, last_row: int) -> None:
    """If the sheet has an Excel table over the tracker, grow it to fit."""
    if last_row <= header_row:
        return
    for table in getattr(ws, "tables", {}).values():
        start, _ = table.ref.split(":")
        if int("".join(filter(str.isdigit, start))) == header_row:
            last_col = get_column_letter(ws.max_column)
            table.ref = f"{start}:{last_col}{last_row}"


def _load_template_workbook(template_path: str | Path) -> Workbook:
    """Load the user's .xlsx template, bounded against a zip bomb.

    Validates the archive's total decompressed size with a bounded read so a
    malicious or corrupt template — tiny on disk, enormous when expanded — is
    rejected with a clear error instead of exhausting memory inside openpyxl.
    Raises ``ValueError`` on a corrupt/invalid workbook or one that exceeds
    :data:`_MAX_TEMPLATE_UNCOMPRESSED_BYTES` when decompressed.
    """
    path = Path(template_path)
    try:
        with zipfile.ZipFile(path) as zf:
            budget = _MAX_TEMPLATE_UNCOMPRESSED_BYTES
            for info in zf.infolist():
                if info.is_dir():
                    continue
                # Fast, clear rejection for a bomb that honestly declares its
                # size; checked against the remaining whole-archive budget.
                if info.file_size > budget:
                    raise ValueError(
                        f"'{path.name}' is too large to process: it declares "
                        f"more than {_MAX_TEMPLATE_UNCOMPRESSED_BYTES:,} bytes of "
                        "decompressed content. Check that this is a real tender "
                        "template."
                    )
                # Bounded read also defends against a header that under-reports
                # the true expanded size; we never hold more than one chunk.
                with zf.open(info) as member:
                    while True:
                        chunk = member.read(_TEMPLATE_READ_CHUNK)
                        if not chunk:
                            break
                        budget -= len(chunk)
                        if budget < 0:
                            raise ValueError(
                                f"'{path.name}' expands to more than "
                                f"{_MAX_TEMPLATE_UNCOMPRESSED_BYTES:,} bytes and "
                                "was refused as a possible zip bomb."
                            )
    except zipfile.BadZipFile as exc:
        raise ValueError(
            f"'{path.name}' is not a valid .xlsx file (corrupt, or not an Excel "
            "workbook). If it's an .xls, re-save it as .xlsx."
        ) from exc
    return load_workbook(str(path))


def create_blank_template() -> Workbook:
    """Build a workbook matching the TENDER_TEMPLATE.xlsx layout."""
    wb = Workbook()
    title_font = Font(bold=True, size=14)
    label_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    ws = wb.active
    ws.title = INFO_SHEET
    ws["A1"] = "TENDER INFORMATION"
    ws["A1"].font = title_font
    for i, label in enumerate(TenderInfo.TEMPLATE_LABELS, start=3):
        cell = ws.cell(row=i, column=1, value=label)
        cell.font = label_font
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 60

    ws = wb.create_sheet(TRACKER_SHEET)
    ws["A1"] = "IT PROCUREMENT COMPLIANCE TRACKER"
    ws["A1"].font = title_font
    for i, header in enumerate(TRACKER_HEADERS, start=1):
        cell = ws.cell(row=3, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill
    widths = [6, 8, 28, 80, 6, 16, 30]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    return wb
